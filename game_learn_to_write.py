import pygame
import sys
import os
import torch
import clip
from PIL import Image
import openpyxl
import pytesseract
from PIL import Image
import Levenshtein
from prediction import prediction

char_recognition = prediction()

def update_excel_file(letter, score, roll_number):
    excel_path = f"{roll_number}_letter_scores.xlsx"

    # Load the existing Excel file
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['Scores']

    # Find the column index for the current letter
    column_values = [col[0].value for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column)]
    
    if letter in column_values:
        column_index = column_values.index(letter) + 1

        # Update the score in the second row for the particular column
        sheet.cell(row=2, column=column_index, value=sheet.cell(row=3, column=column_index).value + score)

        # Increase the value in the third row by 1 for the particular column
        sheet.cell(row=3, column=column_index, value=sheet.cell(row=3, column=column_index).value + 1)

        # Save the changes to the Excel file
        wb.save(excel_path)
    else:
        print(f"Error: The letter '{letter}' is not present in the Excel file.")

# Initialize Pygame
pygame.init()

# Set up display
WIDTH, HEIGHT = 800, 600
screen = pygame.display.set_mode((WIDTH, HEIGHT))
run = True
drawing_color = (35, 35, 35)
screen.fill([200, 205, 255])
dot_color = [0, 0, 0]
dot_radius = 5
add = [325, 225]
dot_position_A = [[28, 200], [40, 156], [80, 55], [124, 156], [138, 200], [80, 156]]

dot_position_B = [[36, 44], [36, 110], [36, 180], [83, 180], [108, 148], [87, 107], [84, 44], [103, 75]]

dot_position_E = [
    [110, 180],
    [73, 180],
    [36, 180],
    [36, 105],
    [36, 105],
    [72, 105],
    [36, 68],
    [110, 105],
    [36, 36],
    [73, 36],
    [110, 36],
    [36, 140],
]
dot_position_F = [[36, 180], [36, 105], [36, 105], [72, 105], [36, 68], [110, 105], [36, 36], [73, 36], [110, 36],
                  [36, 140]]
dot_position_H = [[110, 180], [36, 180], [36, 105], [36, 105], [72, 105], [36, 68], [110, 105], [36, 36], [110, 36],
                  [36, 140], [110, 70], [110, 140]]

dot_position_I = [[110, 36], [36, 36], [110, 180], [36, 180], [72, 36], [72, 105], [72, 180]]

dot_position_L = [[36, 180], [36, 105], [36, 36], [110, 180], [36, 180]]
dot_position_P = [[118, 68], [36, 180], [36, 105], [36, 105], [72, 105], [36, 68], [110, 105], [36, 36], [73, 36],
                  [110, 36], [36, 140]]
dot_position_R = [[118, 68], [36, 180], [36, 105], [36, 105], [72, 105], [36, 68], [110, 105], [36, 36], [73, 36],
                  [110, 36], [36, 140], [110, 180], [75, 146]]

dot_position_T = [[110, 36], [36, 36], [72, 36], [72, 105], [72, 180]]

dot_position_X = [[110, 36], [36, 36], [110, 180], [36, 180], [72, 105], [110, 180], [36, 180]]

dot_position_Y = [[110, 36], [36, 36], [72, 105], [72, 180]]
all_dot_positions = [dot_position_A,  dot_position_B, dot_position_E, dot_position_F, dot_position_H,
                     dot_position_I, dot_position_L, dot_position_P, dot_position_R, dot_position_T, dot_position_X,
                     dot_position_Y]
current_letter=["A", "B", "E", "F", "H", "I", "L", "P", "R", "T", "X", "Y"]
current_dot_positions_index = 0  # Initialize current_dot_positions_index

def calculate_cer(hypothesis, reference):
    return Levenshtein.distance(reference, hypothesis) / len(reference)

def perform_ocr(image_path,letter):
    # Perform OCR on the image
    text = pytesseract.image_to_string(Image.open(image_path), lang='eng')

    # Get ground truth (you need to provide the actual text for comparison)
    reference_text = letter

    # Calculate Character Error Rate (CER)
    cer = calculate_cer(text, reference_text)

    # Get word recognition accuracy
    reference_words = reference_text.split()
    recognized_words = text.split()
    correct_words = [w for w in recognized_words if w in reference_words]
    word_accuracy = len(correct_words) / len(reference_words)

    # Get word confidence levels
    word_confidences = pytesseract.image_to_data(Image.open(image_path), lang='eng', output_type=pytesseract.Output.DICT)['conf']

    
    return word_confidences[0]


def draw_dots(dot_positions):
    for i in dot_positions:
        i[0] = i[0] + add[0]
        i[1] = i[1] + add[1]
        pygame.draw.circle(screen, dot_color, i, dot_radius)
        i[0] = i[0] - add[0]
        i[1] = i[1] - add[1]


def normalize_data_list(data_list):
    if not data_list:
        print("Data list is empty.")
        return None

    # Extract all values from the first element of each inner list
    values = [data[0] for data in data_list if data[0] is not None]

    # Check if all values are zero
    if all(value == 0 for value in values):
        # Handle the case where all values are zero (set to a default value, e.g., 0.5)
        for i in range(len(data_list)):
            data_list[i][0] = 0.5
    else:
        # Normalize the first element of each inner list between 0 and 1
        max_value = max(values)
        min_value = min(values)

        for i in range(len(data_list)):
            if data_list[i][0] is not None:
                # Check for a zero range
                if max_value != min_value:
                    data_list[i][0] = (data_list[i][0] - min_value) / (max_value - min_value)
                else:
                    # Handle the case where the range is zero (set to a default value, e.g., 0.5)
                    data_list[i][0] = 0.5

    return data_list



def extract_data_for_all_letters(roll_number, letters):
    excel_path = f"{roll_number}_letter_scores.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: Excel file '{excel_path}' does not exist.")
        return None

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['Scores']

    data_list = []

    for letter in letters:
        # Find the column index for the specified letter
        column_index = [col[0].value for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column)].index(letter) + 1

        # Extract data from the 2nd and 3rd rows for the specified column
        data_row_2 = sheet.cell(row=2, column=column_index).value
        data_row_3 = sheet.cell(row=3, column=column_index).value

        # Append the data for the current letter to the list
        data_list.append([data_row_2, data_row_3])
    normalize_data_list(data_list)
    factor_of_alph=[]
    index=0
    for item in data_list:
        if(item[0]*item[1]==0):
            factor_of_alph.append(100)
        else:
            factor_of_alph.append(1/(item[0]*item[1]))
    max_index = factor_of_alph.index(max(factor_of_alph))
    return max_index

    return data_array
def next_button_click():
        # Save screenshot of the region (351, 251) to (435, 415)
    letters=["A", "B", "E", "F", "H", "I", "L", "P", "R", "T", "X", "Y"]
    global current_dot_positions_index
    alph=current_dot_positions_index
    screenshot_rect = pygame.Rect(341, 241, 455 - 341, 415 - 241)
    screenshot = pygame.Surface(screenshot_rect.size)
    screenshot.blit(screen, (0, 0), screenshot_rect)
    pygame.image.save(screenshot, "screenshot.png")
    ############
    # score=perform_ocr("screenshot.png",letters[alph])
    score=char_recognition.perform_character_recognition()
    current_dot_positions_index = extract_data_for_all_letters(roll_number, letters)
    screen.fill([200, 205, 255])  # Clear the screen
    draw_dots(all_dot_positions[current_dot_positions_index])
    update_excel_file(letters[alph],score,roll_number)


roll_number = input("Enter your roll number: ")
def initialize_excel_file(letters, roll_number):
    excel_path = f"{roll_number}_letter_scores.xlsx"
    if not os.path.exists(excel_path):
        # Create a new Excel file with a sheet named 'Scores'
        wb = openpyxl.Workbook()
        wb.active.title = 'Scores'
        wb.save(excel_path)

        # Initialize columns with given letters and set values in the second and third rows to 0
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb['Scores']
        for col_num, letter in enumerate(letters, start=1):
            sheet.cell(row=1, column=col_num, value=letter)
            sheet.cell(row=2, column=col_num, value=0)
            sheet.cell(row=3, column=col_num, value=0)
        wb.save(excel_path)
letter=["A", "B", "E", "F", "H", "I", "L", "P", "R", "T", "X", "Y"]
initialize_excel_file(letter,roll_number)
# Main game loop
while run:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            run = False
        elif event.type == pygame.MOUSEMOTION and event.buttons[0]:
            pos = pygame.mouse.get_pos()
            pygame.draw.circle(screen, drawing_color, (pos[0], pos[1]), 5)
            
        elif event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
            # Check if the "Next" button is clicked
            if WIDTH - 150 <= event.pos[0] <= WIDTH and HEIGHT - 100 <= event.pos[1] <= HEIGHT:
                next_button_click()

    # Draw the "Next" button
    pygame.draw.rect(screen, (139, 69, 19), (WIDTH - 150, HEIGHT - 100, 150, 50))  # Brown color
    font = pygame.font.SysFont(None, 36)
    text = font.render("Next", True, (255, 255, 255))  # White text
    screen.blit(text, (WIDTH - 120, HEIGHT - 90))

    pygame.display.update()

pygame.quit()
sys.exit()